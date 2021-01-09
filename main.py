import shelve
import requests
import openpyxl as opx
from datetime import datetime, timedelta


def get_token(login: str, password: str) -> str:
    '''
    Get Token for requests
    :param login:
    :param password:
    :return: token
    '''
    result = requests.post(
        'https://zont-online.ru/api/get_authtoken',
        auth=(login, password),
        headers={'X-ZONT-Client': 'your@email.com'},
        json={"client_name": "My cool app"}
    ).json()
    return result['token']


def get_devices(user_token) -> dict:
    '''
    get all devices and all params
    :return: dict of params
    '''
    url = 'https://zont-online.ru/api/devices'
    data = {'load_io': True}

    headers = {'X-ZONT-Client': 'your@email.com',
               'X-ZONT-Token': user_token
               }
    return requests.post(url, json=data, headers=headers).json()


def write_data_to_excel(data: dict):
    '''
    write data to excel. Create new excel file if not exist, or insert data in exist file (zont.xlsx)
    :param data: dict with params
    :return:
    '''
    boiler_data = data['devices'][0]['io']['last-boiler-state']
    thermometers_data = data['devices'][0]['thermometers']

    try:
        # try to open file. If not exist create it
        wb = opx.load_workbook('zont.xlsx')
        boiler_sheet = wb['ZontIO']
        thermometers_sheet = wb['TermData']

        # calc count of rows in sheet
        boiler_insert_row_number = boiler_sheet.max_row + 1
        term_insert_row_number = thermometers_sheet.max_row + 1
    except:
        # create workbook and sheets
        wb = opx.Workbook()
        boiler_sheet = wb.create_sheet('ZontIO')
        thermometers_sheet = wb.create_sheet('TermData')
        boiler_insert_row_number = 2
        term_insert_row_number = 2

        # insert captions of boiler and term data
        boiler_sheet['A' + '1'] = 'время работы котла'
        boiler_sheet['B' + '1'] = 'авария'
        boiler_sheet['C' + '1'] = 'фактическая температура воды'
        boiler_sheet['D' + '1'] = 'расчётная температура теплоносителя отопления'
        boiler_sheet['E' + '1'] = 'фактическая температура ГВС'
        boiler_sheet['F' + '1'] = 'давление воды (бар)'
        boiler_sheet['G' + '1'] = 'внешнее питание'
        boiler_sheet['H' + '1'] = 'время получения статуса'
        boiler_sheet['I' + '1'] = 'Дата статуса'

        thermometers_sheet['A' + '1'] = 'Имя датчика'
        thermometers_sheet['B' + '1'] = 'Последний статус'
        thermometers_sheet['C' + '1'] = 'Последнее значение'
        thermometers_sheet['D' + '1'] = 'Время'
        thermometers_sheet['E' + '1'] = 'Дата'

    finally:
        # insert boiler and term data
        boiler_sheet['A' + str(boiler_insert_row_number)] = boiler_data['boiler_work_time']
        boiler_sheet['B' + str(boiler_insert_row_number)] = boiler_data['fail']
        boiler_sheet['C' + str(boiler_insert_row_number)] = boiler_data['ot']['bt']
        boiler_sheet['D' + str(boiler_insert_row_number)] = boiler_data['ot']['cs']
        boiler_sheet['E' + str(boiler_insert_row_number)] = boiler_data['ot']['dt']
        boiler_sheet['F' + str(boiler_insert_row_number)] = boiler_data['ot']['wp']
        boiler_sheet['G' + str(boiler_insert_row_number)] = boiler_data['power']
        boiler_sheet['H' + str(boiler_insert_row_number)] = timedelta(seconds=boiler_data['time'] + 10800)
        boiler_sheet['H' + str(boiler_insert_row_number)].number_format = 'hh:mm:ss'
        boiler_sheet['I' + str(boiler_insert_row_number)].number_format = 'DD.MM.YYYY'
        boiler_sheet['I' + str(boiler_insert_row_number)] = datetime.now()

        for term_item in thermometers_data:
            thermometers_sheet['A' + str(term_insert_row_number)] = term_item['name']
            thermometers_sheet['B' + str(term_insert_row_number)] = term_item['last_state']
            thermometers_sheet['C' + str(term_insert_row_number)] = term_item['last_value']
            thermometers_sheet['D' + str(term_insert_row_number)] = timedelta(
                seconds=term_item['last_value_time'] + 10800)
            thermometers_sheet['D' + str(term_insert_row_number)].number_format = 'hh:mm:ss'
            thermometers_sheet['E' + str(term_insert_row_number)] = datetime.now()
            thermometers_sheet['E' + str(term_insert_row_number)].number_format = 'DD.MM.YYYY'
            term_insert_row_number += 1

        wb.save('zont.xlsx')


def auth_user() -> str:
    '''
        authorization user by token or login and password(for get token)
        Read token from shelve. If not exist, retrieve token from API by login and password. Write retrieved token
        into shelve
    :return: token
    '''
    user_choice = 3
    db = shelve.open('token')
    token = None
    try:
        token = db['token']
    except:
        while user_choice != 10:
            print('1 - input Token if you already have, 2 - get token by login password, 0 - exit \n')
            try:
                user_choice = int(input('Enter a choice and press enter: '))
                if user_choice == 0:
                    token = None
                    user_choice = 10
                elif user_choice == 1:
                    token = input('Input Token >> ')
                    user_choice = 10
                elif user_choice == 2:
                    token = get_token(input('Input login>> '), input('Input password>> '))
                    user_choice = 10
            except:
                print('Opps, try one more')
    finally:
        if token:
            db['token'] = token
        db.close()
        return token


if __name__ == '__main__':
    token = auth_user()
    write_data_to_excel(get_devices(token))
